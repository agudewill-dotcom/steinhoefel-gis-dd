import pandas as pd
from io import BytesIO
import streamlit as st

def export_to_excel(df: pd.DataFrame, sheet_name: str = 'Sheet1') -> bytes:
    """Export DataFrame to Excel bytes with formatting."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Formatting
        from openpyxl.styles import PatternFill, Font
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Auto-size columns
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
            
        # Color coding if severity column exists
        if 'severity' in df.columns:
            critical_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Red
            high_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # Orange/Yellow
            medium_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # Light Yellow
            
            sev_idx = df.columns.get_loc('severity') + 1
            for row in range(2, len(df) + 2):
                sev_val = worksheet.cell(row=row, column=sev_idx).value
                fill = None
                if sev_val == 'Critical':
                    fill = critical_fill
                elif sev_val == 'High':
                    fill = high_fill
                elif sev_val == 'Medium':
                    fill = medium_fill
                    
                if fill:
                    for col in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = fill

    return output.getvalue()

def export_to_csv(df: pd.DataFrame) -> bytes:
    """Export to CSV with UTF-8 BOM for Excel compatibility."""
    return df.to_csv(index=False, sep=';').encode('utf-8-sig')

def render_download_buttons(df: pd.DataFrame, name: str, key_prefix: str):
    """Render Streamlit download buttons for Excel and CSV."""
    col1, col2 = st.columns([1,1])
    with col1:
        st.download_button(
            label=f"Download {name} (Excel)",
            data=export_to_excel(df),
            file_name=f"{name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key_prefix}_excel"
        )
    with col2:
        st.download_button(
            label=f"Download {name} (CSV)",
            data=export_to_csv(df),
            file_name=f"{name}.csv",
            mime="text/csv",
            key=f"{key_prefix}_csv"
        )
